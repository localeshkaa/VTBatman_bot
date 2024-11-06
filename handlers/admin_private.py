from aiogram import F, Router, types
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from openpyxl.reader.excel import load_workbook

from edit_table import get_name, get_group, delete_person_from_xlsx, create_file_buttons, save_file_buttons, \
    process_user_input
from filters.chat_types import ChatTypeFilter, IsAdmin
from kbds.reply import get_keyboard
from states import EditTable

admin_router = Router()
admin_router.message.filter(ChatTypeFilter(["private"]), IsAdmin())

ADMIN_KB = get_keyboard(
    "Создать файл",
    "Отобразить таблицу",
    "Добавить человека",
    "Удалить человека",
    "Создать группу",
    "Закрыть и сохранить сегодняшний файл",
    "Назад",
    placeholder="Выберите действие",
    sizes=(2,)
)

EXT_KB = get_keyboard(
    "Назад",
    placeholder=None,
    sizes=(2,)
)


@admin_router.message(Command("admin"))
async def admin_features(message: types.Message):
    await message.answer("Что хотите сделать?", reply_markup=ADMIN_KB)


@admin_router.message(StateFilter(None), F.text == "Создать файл")
async def create_file(message: types.Message, state: FSMContext):
    await message.answer("""
    Введите исходные данные по примеру ниже:
    КК ЦО (план): 
    АС ЦО (план):
    СОМ ЦО (план):
    АП (ЖКУ) план:
    """, reply_markup=EXT_KB)
    await create_file_buttons(message, state)  # Передаем сообщение для дальнейшей обработки
    await state.set_state(EditTable.creating_file)


@admin_router.message(EditTable.creating_file)
async def set_constants(message: types.Message, state: FSMContext):
    await message.answer("Файл был создан", reply_markup=EXT_KB)
    await process_user_input(message, state)
    await state.clear()


@admin_router.message(EditTable.creating_file, F.text == "Назад")
async def cancel_creation(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Создание файла отменено. Выберите действия:", reply_markup=ADMIN_KB)


@admin_router.message(StateFilter(None), F.text == "Закрыть и сохранить сегодняшний файл")
async def save_file(message: types.Message):
    await save_file_buttons(message)
    await message.answer("Файл сохранен и закрыт", reply_markup=ADMIN_KB)


@admin_router.message(StateFilter(None), F.text == "Добавить человека")
async def start_add_person(message: types.Message, state: FSMContext):
    await message.answer("Введите фамилию и имя человека", reply_markup=EXT_KB)
    await state.set_state(EditTable.add_name)


@admin_router.message(EditTable.add_name, F.text == "Назад")
async def cancel_add_person(message: types.Message, state: FSMContext):
    await state.clear()  # Сброс состояния
    await message.answer("Добавление человека отменено. Выберите действие:", reply_markup=ADMIN_KB)


@admin_router.message(EditTable.add_name)
async def add_person_name(message: types.Message, state: FSMContext):
    # Сохраняем имя и переходим к запросу группы
    await state.update_data(name=message.text)
    await message.answer("Введите номер группы (например, 1 для 'Группа 1')")
    await state.set_state(EditTable.add_group_id)


@admin_router.message(EditTable.add_group_id, F.text == "Назад")
async def cancel_add_group(message: types.Message, state: FSMContext):
    await state.clear()  # Сброс состояния
    await message.answer("Добавление группы отменено. Выберите действие:", reply_markup=ADMIN_KB)


@admin_router.message(EditTable.add_group_id)
async def add_person_group(message: types.Message, state: FSMContext):
    await state.update_data(group_id=message.text)

    # Вызов функции get_group для добавления в таблицу
    try:
        await get_group(message, state)  # Убедитесь, что get_group принимает только message и state
        await message.answer("Человек добавлен в таблицу", reply_markup=ADMIN_KB)
    except Exception as e:
        await message.answer(f"Произошла ошибка: {str(e)}")

    await state.clear()  # Ensure state is cleared after adding


@admin_router.message(StateFilter(None), F.text == "Удалить человека")
async def start_delete_person(message: types.Message, state: FSMContext):
    await message.answer("Введите фамилию и имя человека, которого хотите удалить", reply_markup=EXT_KB)
    await state.set_state(EditTable.delete_name)


@admin_router.message(EditTable.delete_name)
async def delete_person_name(message: types.Message, state: FSMContext):
    # Сохраняем имя для удаления
    name_to_delete = message.text
    await state.update_data(delete_name=name_to_delete)

    try:
        workbook = load_workbook("shablon.xlsx")  # Make sure you load the workbook here
        sheet_found = False
        for sheet in workbook.sheetnames:
            current_sheet = workbook[sheet]
            result = await delete_person_from_xlsx(current_sheet, name_to_delete)
            if result:
                sheet_found = True
                break

        if sheet_found:
            workbook.save("shablon.xlsx")
            await message.answer("Человек удален из таблицы", reply_markup=ADMIN_KB)
        else:
            await message.answer("Человек не найден или не может быть удален", reply_markup=ADMIN_KB)

        workbook.close()  # Always close the workbook after modification
    except Exception as e:
        await message.answer(f"Произошла ошибка: {str(e)}")

    await state.clear()  # Ensure state is cleared after deleting


@admin_router.message(EditTable.show_data, F.text == "Отобразить таблицу")
async def show_data(message: types.Message, state: FSMContext):
    await state.update_data(show_data=message.text)
    await message.answer("Введите номер группы")
    await message.answer("Функция в разработке", reply_markup=ADMIN_KB)
    await state.clear()  # Clear the state at the end of the process
