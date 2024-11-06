import datetime
import re

from openpyxl import load_workbook
from aiogram import types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.fsm.context import FSMContext

from states import EditTable

current_file = None


async def get_name(name_message: types.Message, state: FSMContext):
    name = name_message.text.strip()
    await state.update_data(name=name)  # Сохраняем имя в состоянии
    await name_message.answer("Введите номер группы (например, 1 для 'Группа 1')", reply_markup=get_back_markup())
    await state.set_state(EditTable.add_group_id)


async def get_group(group_message: types.Message, state: FSMContext):
    data = await state.get_data()
    name = data.get('name')
    group_id = group_message.text.strip()
    sheet_name = f"Группа {group_id}"

    try:
        file = load_workbook("shablon.xlsx")
        ws = file[sheet_name] if sheet_name in file.sheetnames else None

        if ws is None:
            await group_message.answer(f"Лист '{sheet_name}' не найден. Проверьте номер группы.")
            return

        # Проверка на наличие имени в листе
        found = any(cell.value == name for cell in ws['A'])

        if found:
            await group_message.answer(f"Имя '{name}' уже присутствует в листе '{sheet_name}'.",
                                       reply_markup=get_back_markup())
            await state.set_state(EditTable.awaiting_back)
            return

        # Добавление имени в первую свободную ячейку
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value is None:
                ws[f'A{row}'] = name
                break
            elif cell_value == "Итог":
                await group_message.answer(f"Не удалось добавить имя '{name}', так как достигнуто слово 'Итог'.")
                return

        file.save("shablon.xlsx")
        await group_message.answer(f"Имя '{name}' добавлено в лист '{sheet_name}'.")

    except Exception as e:
        await group_message.answer(f"Произошла ошибка: {str(e)}")
    finally:
        file.close()


def get_back_markup():
    # Функция для создания клавиатуры с кнопкой "Назад"
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton("Назад")]
        ],
        resize_keyboard=True
    )


# Проверка существования имени в файле
def find_name_in_workbook(file, name):
    for sheet in file.sheetnames:
        ws = file[sheet]
        for cell in ws['A']:
            if cell.value == name:
                return True, sheet
    return False, None


# Добавление имени в лист, если доступно свободное место
async def add_name_to_sheet(sheet, name):
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value is None:
            sheet[f'A{row}'] = name
            return True
        elif cell_value == "Итог":
            return False
    return False


# Отправка сообщения, если имя уже существует
async def send_name_exists_message(message, name, sheet, state):
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton("Назад"))
    await message.answer(f"Имя '{name}' уже присутствует в листе '{sheet}'.", reply_markup=markup)
    await state.set_state(EditTable.awaiting_back)  # Устанавливаем состояние ожидания нажатия "Назад"


async def delete_person_from_xlsx(sheet, name):
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value == name:
            sheet.delete_rows(row, 1)  # Удаляет строку с найденным именем
            return True
        elif cell_value == "Итог":
            return False
    return False


async def create_file_buttons(file_message: types.Message, state: FSMContext):
    global current_file  # Указываем, что переменная глобальная

    # Получаем сегодняшнюю дату
    today_date = datetime.date.today()
    file_name = today_date.strftime("%Y%m%d") + ".xlsx"  # Имя файла с датой и расширением .xlsx

    # Загружаем шаблон и создаем копию
    workbook = load_workbook('shablon.xlsx')
    current_file = file_name  # Сохраняем имя файла для дальнейшего использования

    # Сохраняем новый файл с именем, основанным на сегодняшней дате
    workbook.save(current_file)

    # Сохраняем имя файла в состоянии
    await state.update_data(file_message=file_message)

    # Запрашиваем у пользователя исходные данные

    await state.set_state(EditTable.waiting_for_data)  # Переход к состоянию ожидания ввода данных


# Обработка данных после ввода пользователем
async def process_user_input(message: types.Message, state: FSMContext):
    global current_file  # Используем глобальную переменную для имени текущего файла

    # Получаем текущие данные
    user_input = message.text
    kk_co_plan = re.search(r'КК ЦО (план): (.+)', user_input).group(1) if re.search(r'КК ЦО (план): (.+)',
                                                                                    user_input) else None
    ac_co_plan = re.search(r'АС ЦО (план): (.+)', user_input).group(1) if re.search(r'АС ЦО (план): (.+)',
                                                                                    user_input) else None
    com_co_plan = re.search(r'СОМ ЦО (план): (.+)', user_input).group(1) if re.search(r'СОМ ЦО (план): (.+)',
                                                                                      user_input) else None
    ap_zku_plan = re.search(r'АП (ЖКУ) план: (.+)', user_input).group(1) if re.search(r'АП (ЖКУ) план: (.+)',
                                                                                      user_input) else None

    # Загружаем текущий файл
    workbook = load_workbook(current_file)

    # Обрабатываем все листы в книге
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]

        # Проходим по всем строкам начиная с 2-й (предполагаем, что 1-я строка - заголовки)
        for row in current_sheet.iter_rows(min_row=2, values_only=False):  # Начинаем с 2-й строки
            # Проверяем, что в первой ячейке есть значение, отличное от None и "Итог"
            if row[0].value not in [None, "Итог"]:
                # Записываем данные в соответствующие столбцы
                row[1].value = kk_co_plan if kk_co_plan else row[1].value
                row[2].value = ac_co_plan if ac_co_plan else row[2].value
                row[3].value = com_co_plan if com_co_plan else row[3].value
                row[4].value = ap_zku_plan if ap_zku_plan else row[4].value

    # Сохраняем изменения в файл
    try:
        workbook.save(current_file)
    except Exception as e:
        await message.answer(f"Ошибка при сохранении файла: {str(e)}")

    # Подтверждаем успешное выполнение
    await message.answer(f"Данные обновлены во всех листах созданного файла '{current_file}'.")
    await state.set_state(EditTable.file_created)  # Переход в состояние "файл создан"


async def save_file_buttons(message: types.Message):
    global current_file
    if current_file:
        workbook = load_workbook(current_file)
        workbook.save(current_file)
        workbook.close()
        await message.answer(f"Файл '{current_file}' сохранен и закрыт.")
        current_file = None  # Сбрасываем переменную после сохранения
    else:
        await message.answer("Файл не найден. Создайте файл перед сохранением.")


async def add_note_xlsx(message: types.Message):
    global current_file
    if not current_file:
        await message.answer("Сначала администратор должен создать файл.")
        return

    workbook = load_workbook(current_file)
    # Получаем активный лист, можно использовать find_name_in_workbook для поиска по имени листа

    user_input = message.text
    name_match = re.search(r'ФИО: (.+)', user_input)
    if not name_match:
        await message.answer("Ошибка: ФИО не найдено в вашем сообщении.")
        return

    name = name_match.group(1)
    sheet = find_name_in_workbook(current_file, name)

    # Извлекаем все данные из сообщения с помощью регулярных выражений
    kk_co_fact = re.search(r'КК (факт): (.+)', user_input).group(1) if re.search(r'КК (факт): (.+)',
                                                                                 user_input) else None
    as_co_fact = re.search(r'АС(факт): (.+)', user_input).group(1) if re.search(r'АС(факт): (.+)', user_input) else None
    com_co_fact = re.search(r'СОМ ЦО (факт): (.+)', user_input).group(1) if re.search(r'СОМ ЦО (факт): (.+)',
                                                                                      user_input) else None
    ap_zku_co_fact = re.search(r'АП (ЖКУ) ЦО факт: (.+)', user_input).group(1) if re.search(r'АП (ЖКУ) ЦО факт: (.+)',
                                                                                            user_input) else None
    pens = re.search(r'ПЕНС: (.+)', user_input).group(1) if re.search(r'ПЕНС: (.+)', user_input) else None
    izp = re.search(r'ИЗП: (.+)', user_input).group(1) if re.search(r'ИЗП: (.+)', user_input) else None

    # Проходим по всем строкам начиная с 2-й, чтобы найти строку с нужным ФИО
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == name:  # Сравнение с ФИО в первом столбце
            # Обновляем значения в соответствующих столбцах
            row[2].value = kk_co_fact if kk_co_fact else row[1].value
            row[5].value = as_co_fact if as_co_fact else row[2].value
            row[8].value = com_co_fact if com_co_fact else row[3].value
            row[11].value = ap_zku_co_fact if ap_zku_co_fact else row[4].value
            row[14].value = pens if pens else row[5].value
            row[15].value = izp if izp else row[6].value

            # Сохраняем изменения в файле
            workbook.save(current_file)
            workbook.close()

            # Отправляем сообщение о успешном обновлении данных
            await message.answer(f"Данные для '{name}' обновлены.")
            break
    else:
        await message.answer(f"Запись с ФИО '{name}' не найдена в файле.")
        workbook.close()  # Закрываем файл при отсутствии изменений
